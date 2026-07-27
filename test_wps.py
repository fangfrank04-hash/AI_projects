import asyncio
from pathlib import Path
from openpyxl import load_workbook
# 导入tool.py里的wps_to_file函数，路径根据你项目结构调整
from app.utils.tool import wps_to_file

# ========== 你只需要改这里的配置 ==========
# 1. 你的Excel文件完整路径（内网本地路径）
EXCEL_PATH = r"D:\你的文件夹\数据表格.xlsx"
# 2. PDF输出的文件夹（不存在会自动创建）
OUTPUT_DIR = r"D:\你的文件夹\输出结果"
# 3. HTML内容的列名（和你表格里一致，不用改）
HTML_COLUMN_NAME = "info_body"
# ==========================================

async def main():
    # 第一步：读取Excel，拼接所有HTML片段
    print("正在读取Excel并拼接HTML...")
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active

    # 找到info_body列的位置
    headers = [cell.value for cell in ws[1]]
    html_col_idx = headers.index(HTML_COLUMN_NAME) + 1

    # 按行拼接成完整HTML
    html_parts = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        html_content = row[html_col_idx - 1]
        if html_content:
            html_parts.append(str(html_content))

    # 包上完整的html外壳
    full_html = "<html><body>" + "".join(html_parts) + "</body></html>"

    # 第二步：把拼接好的HTML存成本地临时文件
    temp_html_path = Path(OUTPUT_DIR) / "temp_full.html"
    temp_html_path.parent.mkdir(parents=True, exist_ok=True)
    temp_html_path.write_text(full_html, encoding="utf-8")
    print(f"HTML拼接完成，临时文件已保存到: {temp_html_path}")

    # 第三步：调用WPS接口转换PDF
    print("正在调用WPS接口转换PDF...")
    try:
        pdf_path = await wps_to_file(
            input_path=temp_html_path,
            output_dir=OUTPUT_DIR,
            convert="pdf"
        )
        print(f"✅ 转换成功！PDF文件在: {pdf_path}")
    except Exception as e:
        print(f"❌ 转换失败: {e}")

if __name__ == "__main__":
    asyncio.run(main())